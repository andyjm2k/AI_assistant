import base64
from io import BytesIO
from pathlib import Path
from typing import Any, Collection, Dict, List, Optional

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from PIL import Image
except ImportError:
    Image = None


MISSING_FILE_READER_DEPENDENCIES: List[str] = []
if Document is None:
    MISSING_FILE_READER_DEPENDENCIES.append("python-docx")
if openpyxl is None:
    MISSING_FILE_READER_DEPENDENCIES.append("openpyxl")
if PyPDF2 is None:
    MISSING_FILE_READER_DEPENDENCIES.append("PyPDF2")
if Image is None:
    MISSING_FILE_READER_DEPENDENCIES.append("Pillow")

FILE_READERS_AVAILABLE = not MISSING_FILE_READER_DEPENDENCIES


def _require_dependency(dependency: Optional[object], package_name: str) -> None:
    if dependency is not None:
        return
    raise RuntimeError(f"{package_name} is required for this file type.")


def read_text_file(filepath: Path) -> str:
    """Read a plain text file and return its content."""
    try:
        with filepath.open("r", encoding="utf-8") as handle:
            return handle.read()
    except UnicodeDecodeError:
        with filepath.open("r", encoding="latin-1") as handle:
            return handle.read()


def read_docx_file(filepath: Path) -> str:
    """Read a Word document and return its text content."""
    _require_dependency(Document, "python-docx")
    doc = Document(filepath)
    paragraphs = [para.text for para in doc.paragraphs]
    return "\n".join(paragraphs)


def read_xlsx_file(filepath: Path) -> str:
    """Read an Excel file and return its content as formatted text."""
    _require_dependency(openpyxl, "openpyxl")
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    result: List[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        result.append(f"=== Sheet: {sheet_name} ===\n")
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            result.append("\t".join(row_data))
        result.append("\n")

    return "\n".join(result)


def read_pdf_file(filepath: Path) -> str:
    """Read a PDF file and return its text content."""
    _require_dependency(PyPDF2, "PyPDF2")
    result: List[str] = []
    with filepath.open("rb") as handle:
        pdf_reader = PyPDF2.PdfReader(handle)
        for page_num, page in enumerate(pdf_reader.pages, start=1):
            text = page.extract_text()
            result.append(f"=== Page {page_num} ===\n{text}\n")
    return "\n".join(result)


def read_png_file(filepath: Path) -> Dict[str, Any]:
    """Read a PNG image and return metadata and base64-encoded data."""
    _require_dependency(Image, "Pillow")
    img = Image.open(filepath)
    metadata = {
        "width": img.width,
        "height": img.height,
        "format": img.format,
        "mode": img.mode,
    }

    buffered = BytesIO()
    img.save(buffered, format=img.format)
    img_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

    return {
        "metadata": metadata,
        "data": img_base64,
        "description": f"Image: {img.width}x{img.height} pixels, format: {img.format}",
    }


def read_supported_file_text(filepath: Path, text_file_extensions: Collection[str]) -> tuple[str, str]:
    ext = filepath.suffix.lower()
    if ext in text_file_extensions:
        return read_text_file(filepath), "text"
    if ext == ".docx":
        return read_docx_file(filepath), "text"
    if ext in {".xlsx", ".xls"}:
        return read_xlsx_file(filepath), "text"
    if ext == ".pdf":
        return read_pdf_file(filepath), "text"
    raise ValueError(f"Unsupported text-readable file type: {ext}")


__all__ = [
    "FILE_READERS_AVAILABLE",
    "MISSING_FILE_READER_DEPENDENCIES",
    "read_docx_file",
    "read_pdf_file",
    "read_png_file",
    "read_supported_file_text",
    "read_text_file",
    "read_xlsx_file",
]
